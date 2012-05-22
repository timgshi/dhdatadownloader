from django.core.management import setup_environ
import settings
setup_environ(settings)

from Frameworks import ParsePy
from main.models import DHPhoto, DHUserProfile
from django.db.models import Max
import datetime

ParsePy.APPLICATION_ID = "53Rdo20D9PA1hiPTN7qPzcPVaNQEmAkMXi3j6tLv"
ParsePy.MASTER_KEY = "FqhBINgpfI1ISF1ao2poRHYzvhbbr6PjJvuij0cq"

def resetDB():
    print 'reset db start'
    total = 0
    while True:
        print '===NEW LOOP==='
        query = ParsePy.ParseQuery("DHPhoto").limit(100).skip(total)
        query.order("createdAt", False)
        print "Fetching......."
        objects = query.fetch();
        print "....done"
        count = 0
        for x in objects:
            print x.createdAt()
            try:
                existingPhoto = DHPhoto.objects.get(objectID=x.objectId())
                existingPhoto.params = x.__dict__
                existingPhoto.createdAt = x.createdAt()
                existingPhoto.save()
                print 'updated existing'
            except DHPhoto.DoesNotExist:
                photo = DHPhoto(objectID=x.objectId(), createdAt=x.createdAt(), params=x.__dict__)
                photo.save()
                print 'new created'
            count += 1
            print total + count
        total += count
        print total
        if count == 0:
            break

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

from openpyxl.cell import get_column_letter

def writeFile():

	now = datetime.datetime.now()

	wb = Workbook()

	# dest_filename = r'dhdata-%s.xlsx' % now.strftime('%d-%m-%y')
	dest_filename = r'dhdata-latest.xlsx'
	ws = wb.worksheets[0]

	ws.title = "dh data"

	# for col_idx in xrange(1, 40):
	#     col = get_column_letter(col_idx)
	#     for row in xrange(1, 600):
	#         ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

	objects = DHPhoto.objects.order_by('-createdAt')
	headers = []
	row = 2
	for photo in objects:
	    photoArray = []
	    for x in photo.params:
	        
	        if x not in 'DHDataWhoTook' and x not in 'ACL':
	            col = 0
	            try:
	                col = headers.index(x)
	            except ValueError:
	                headers.append(x)
	                col = headers.index(x)
	            col += 1
	            cellValue = "%s" % photo.params.get(x)
	            cellValue = cellValue.encode('ascii', 'ignore')
	            # photoArray.insert(col, cellValue)
	            # print photoArray
	            colLetter = get_column_letter(col)
	            ws.cell('%s%s'%(colLetter,row)).value = cellValue
	            # ws.write(row, col, "%s" % photo.params.get(x))
	    # csvArray.insert(row, photoArray)
	    row += 1
	print headers
	col = 1
	for header in headers:
		headerLetter = get_column_letter(col)
		ws.cell('%s%s'%(headerLetter, 1)).value = header
		col += 1

	wb.save(filename = dest_filename)

def getSmiles():
	total = 0
	while True:
	        print '===NEW LOOP==='
	        query = ParsePy.ParseQuery("DHPhotoSmile").limit(100).skip(total)
	        query.order("createdAt", False)
	        print "Fetching......."
	        try:
	            objects = query.fetch();
	            print "....done"
	            count = 0
	            for x in objects:
	                try:
	                    print x.createdAt()
	                    photoID = x.DHPhotoID
	                    photo = ParsePy.ParseQuery("DHPhoto").get(photoID)
	                    try:
	                        smilesArray = photo.smiles
	                        smilesArray.append(x)
	                        photo.smiles = smilesArray
	                        photo.smileCount = len(smilesArray)
	                    except AttributeError:
	                        smilesArray = [x]
	                        photo.smiles = smilesArray
	                        photo.smileCount = 1
	                    photo.save()
	                    count += 1
	                except (urllib2.HTTPError, urllib2.URLError):
	                    print "PHOTO COULDN\"T BE SAVED"
	                print total + count
	            total += count
	            print total
	            if count == 0:
	                break
	        except urllib2.URLError:
	            print "URL ERROR"

# resetDB()
writeFile()